#!/usr/bin/env python3
"""Merge the competition water-quality workbooks and write a data audit."""

from __future__ import annotations

import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
OUTPUT_PATH = DATA_DIR / "merged.xlsx"
README_PATH = ROOT / "report" / "preprocess.md"

STANDARD_COLUMNS = [
    "DATE",
    "TIME",
    "RIVER LEVEL",
    "R/W PUMP DUTY",
    "R/W FLOW",
    "R/W NTU",
    "R/W CLR",
    "R/W PH",
    "FILT. NTU",
    "C/W WELL LEVEL",
    "PH",
    "NTU",
    "CLR",
    "CL2",
    "F/RIDE",
    "ALUM",
    "T/W PUMP DUTY",
    "T/W FLOW",
    "18ML LEVEL",
    "18ML FLOW",
    "REMARKS",
]

DROP_COLUMNS = ["18ML LEVEL", "18ML FLOW", "REMARKS"]
OUTPUT_COLUMNS = [
    column for column in STANDARD_COLUMNS if column not in DROP_COLUMNS
]

NUMERIC_COLUMNS = [
    "RIVER LEVEL",
    "R/W FLOW",
    "R/W NTU",
    "R/W CLR",
    "R/W PH",
    "FILT. NTU",
    "C/W WELL LEVEL",
    "PH",
    "NTU",
    "CLR",
    "CL2",
    "F/RIDE",
    "ALUM",
    "T/W FLOW",
]

PUMP_COLUMNS = ["R/W PUMP DUTY", "T/W PUMP DUTY"]
MAX_IMPUTE_GAP = 6

HARD_RANGES = {
    "RIVER LEVEL": (0, 30),
    "R/W FLOW": (0, 100),
    "R/W NTU": (0, None),
    "R/W CLR": (0, None),
    "R/W PH": (0, 14),
    "FILT. NTU": (0, None),
    "C/W WELL LEVEL": (0, 30),
    "PH": (0, 14),
    "NTU": (0, None),
    "CLR": (0, None),
    "CL2": (0, 20),
    "F/RIDE": (0, 1),
    "ALUM": (0, 1),
    "T/W FLOW": (0, 100),
    "18ML LEVEL": (0, 30),
    "18ML FLOW": (0, 100),
}

EXPECTED_TIMES = [
    "0700",
    "0900",
    "1100",
    "1300",
    "1500",
    "1700",
    "1900",
    "2100",
    "2300",
    "0100",
    "0300",
    "0500",
]

MONTH_NAMES = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "July": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

HEADER_ALIASES = {
    "DATA": "DATE",
    "DATE": "DATE",
    "TIME": "TIME",
    "RIVER LEVEL": "RIVER LEVEL",
    "R/W PUMP DUTY": "R/W PUMP DUTY",
    "R/W FLOW": "R/W FLOW",
    "R/W NTU": "R/W NTU",
    "R/W CLR": "R/W CLR",
    "R/W PH": "R/W PH",
    "FILT. NTU": "FILT. NTU",
    "C/W WELL LEVEL": "C/W WELL LEVEL",
    "PH": "PH",
    "NTU": "NTU",
    "CLR": "CLR",
    "CL2": "CL2",
    "F/RIDE": "F/RIDE",
    "ALUM": "ALUM",
    "T/W PUMP DUTY": "T/W PUMP DUTY",
    "T/W FLOW": "T/W FLOW",
    "18ML LEVEL": "18ML LEVEL",
    "18ML FLOW": "18ML FLOW",
    "REMARKS": "REMARKS",
}

DATA_DICTIONARY = [
    ("DATE", "自然日期", "由来源运营日和时点重建"),
    ("TIME", "监测时间", "每天 12 次，每 2 小时一次"),
    ("RIVER LEVEL", "河水水位", "题目未注明单位"),
    ("R/W PUMP DUTY", "原水泵运行数量", "由原泵组编号中的数字数量转换"),
    ("R/W FLOW", "原水流量", "题目未注明单位"),
    ("R/W NTU", "原水浊度", "NTU"),
    ("R/W CLR", "原水色度", "题目未注明单位"),
    ("R/W PH", "原水 pH 值", "无量纲"),
    ("FILT. NTU", "滤后水浊度", "NTU"),
    ("C/W WELL LEVEL", "清水池水位", "题目未注明单位"),
    ("PH", "清水/处理后水 pH 值", "无量纲"),
    ("NTU", "清水/处理后水浊度", "NTU"),
    ("CLR", "清水/处理后水色度", "题目未注明单位"),
    ("CL2", "余氯", "题目未注明单位"),
    ("F/RIDE", "矾/混凝剂投加流量", "题目未注明单位"),
    ("ALUM", "明矾/混凝剂投加量", "题目未注明单位"),
    ("T/W PUMP DUTY", "送水泵运行数量", "由原泵组编号中的数字数量转换"),
    ("T/W FLOW", "送水流量/出厂水流量", "题目未注明单位"),
    ("18ML LEVEL", "1800 万升水池水位", "题目未注明单位"),
    ("18ML FLOW", "1800 万升水池进出水流量", "题目未注明单位"),
    ("REMARKS", "备注", "异常事件或人工操作记录"),
]

README_MARKER = "<!-- DATA_AUDIT_START -->"


def normalize_header(value: Any) -> str | None:
    if pd.isna(value):
        return None
    text = re.sub(r"\s+", " ", str(value).strip().upper())
    return HEADER_ALIASES.get(text)


def is_missing(value: Any) -> bool:
    if pd.isna(value):
        return True
    return isinstance(value, str) and value.strip() in {"", "-", "--", "—"}


def clean_value(value: Any) -> Any:
    if is_missing(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, str):
        return value.strip()
    return value


def convert_pump_duties(
    audit: pd.DataFrame,
) -> tuple[pd.DataFrame, list[str], dict[str, Counter]]:
    """Convert raw/treated-water pump identifiers to running-pump counts."""
    source_counts: dict[str, Counter] = {}
    corrections: list[str] = []

    for column in PUMP_COLUMNS:
        converted: list[int | None] = []
        value_counts: Counter = Counter()
        for index, value in audit[column].items():
            if is_missing(value):
                converted.append(None)
                continue

            text = str(value).strip()
            if text.endswith(".0") and text[:-2].isdigit():
                text = text[:-2]
            value_counts[text] += 1

            if column == "T/W PUMP DUTY" and text == "244":
                converted.append(2)
                corrections.append(
                    "2025-09-18 11:00 的 `T/W PUMP DUTY=244`，上下相邻记录均为 "
                    "`2+4` 或 `2&4`，判定为泵号 2、4 的分隔符误录，转换为运行泵数量 2。"
                )
                continue
            if column == "T/W PUMP DUTY" and text == "2,2":
                converted.append(2)
                corrections.append(
                    "2025-04-08 05:00 的 `T/W PUMP DUTY=2,2`，前后相邻记录均为 "
                    "`2,3`，判定为泵号 2、3 的误录，转换为运行泵数量 2。"
                )
                continue

            pump_ids = re.findall(r"\d+", text)
            if not pump_ids:
                converted.append(None)
                corrections.append(
                    f"{audit.at[index, '_TIMESTAMP']:%Y-%m-%d %H:%M} 的 "
                    f"`{column}={text}` 无法识别，转换为空值。"
                )
                continue
            converted.append(len(pump_ids))

        audit[column] = pd.array(converted, dtype="Int64")
        source_counts[column] = value_counts
    return audit, corrections, source_counts


def normalize_time(value: Any, expected: str) -> tuple[str, str | None]:
    if is_missing(value):
        return expected, f"missing time repaired to {expected}"

    if isinstance(value, (int, np.integer)):
        candidate = f"{int(value):04d}"
    elif isinstance(value, (float, np.floating)) and float(value).is_integer():
        candidate = f"{int(value):04d}"
    else:
        text = str(value).strip()
        candidate = text.zfill(4) if text.isdigit() else text

    if candidate in EXPECTED_TIMES:
        return candidate, None
    return expected, f"invalid time {value!r} repaired to {expected}"


def month_from_2025_filename(path: Path) -> int:
    match = re.search(r"JBALB_([A-Za-z]+)2025", path.stem)
    if not match or match.group(1) not in MONTH_NAMES:
        raise ValueError(f"Cannot infer month from {path.name}")
    return MONTH_NAMES[match.group(1)]


def parse_2025_file(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    month = month_from_2025_filename(path)
    raw = pd.read_excel(path, sheet_name=0, header=None, engine="openpyxl")
    headers = [normalize_header(value) for value in raw.iloc[0].tolist()]
    rows: list[dict[str, Any]] = []
    repairs: list[str] = []
    data = raw.iloc[1:].dropna(how="all").reset_index(drop=True)
    expected_rows = pd.Period(datetime(2025, month, 1), freq="M").days_in_month * 12
    if len(data) != expected_rows:
        raise ValueError(f"{path.name}: expected {expected_rows} rows, found {len(data)}")

    for index, source_row in data.iterrows():
        day = index // 12 + 1
        slot = index % 12
        operational_date = date(2025, month, day)
        expected_time = EXPECTED_TIMES[slot]
        raw_time = source_row.iloc[1]
        time_text, repair = normalize_time(raw_time, expected_time)
        if repair:
            repairs.append(
                f"{path.name}, source row {index + 2}, {repair}"
            )
        hour = int(time_text[:2])
        timestamp_date = operational_date + timedelta(days=1 if hour < 7 else 0)

        record = {column: None for column in STANDARD_COLUMNS}
        record["DATE"] = timestamp_date
        record["TIME"] = time(hour, 0)
        for col_index, header in enumerate(headers):
            if header in {None, "DATE", "TIME"}:
                continue
            record[header] = clean_value(source_row.iloc[col_index])

        record["_SOURCE_FILE"] = path.name
        record["_SOURCE_SHEET"] = "Sheet1"
        record["_SOURCE_DATE"] = operational_date
        record["_SOURCE_ROW"] = index + 2
        rows.append(record)
    return rows, repairs


def parse_2026_file(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    year = int(path.stem[:4])
    workbook = pd.ExcelFile(path, engine="xlrd")
    rows: list[dict[str, Any]] = []
    repairs: list[str] = []

    for sheet_name in workbook.sheet_names:
        match = re.match(r"\s*(\d{1,2})\.(\d{1,2})", sheet_name)
        if not match:
            raise ValueError(f"{path.name}: invalid sheet name {sheet_name!r}")
        day, month = map(int, match.groups())
        operational_date = date(year, month, day)
        raw = pd.read_excel(
            path, sheet_name=sheet_name, header=None, engine="xlrd"
        ).dropna(how="all")
        if len(raw) != 13:
            raise ValueError(
                f"{path.name}/{sheet_name}: expected 13 rows, found {len(raw)}"
            )
        headers = [normalize_header(value) for value in raw.iloc[0].tolist()]
        data = raw.iloc[1:].reset_index(drop=True)

        for index, source_row in data.iterrows():
            expected_time = EXPECTED_TIMES[index]
            raw_time = source_row.iloc[0]
            time_text, repair = normalize_time(raw_time, expected_time)
            if repair:
                repairs.append(
                    f"{path.name}/{sheet_name}, source row {index + 2}, {repair}"
                )
            hour = int(time_text[:2])
            timestamp_date = operational_date + timedelta(days=1 if hour < 7 else 0)

            record = {column: None for column in STANDARD_COLUMNS}
            record["DATE"] = timestamp_date
            record["TIME"] = time(hour, 0)
            for col_index, header in enumerate(headers):
                if header in {None, "DATE", "TIME"}:
                    continue
                record[header] = clean_value(source_row.iloc[col_index])

            record["_SOURCE_FILE"] = path.name
            record["_SOURCE_SHEET"] = sheet_name.strip()
            record["_SOURCE_DATE"] = operational_date
            record["_SOURCE_ROW"] = index + 2
            rows.append(record)
    return rows, repairs


def coerce_numeric_columns(
    audit: pd.DataFrame,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    invalid: list[dict[str, Any]] = []
    for column in NUMERIC_COLUMNS:
        converted = pd.to_numeric(audit[column], errors="coerce")
        bad_mask = audit[column].notna() & converted.isna()
        for index in audit.index[bad_mask]:
            invalid.append(anomaly_record(audit, index, column, audit.at[index, column]))
        audit[column] = converted
    return audit, invalid


def missing_runs(mask: pd.Series) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start: int | None = None
    for index, missing in enumerate(mask.tolist()):
        if missing and start is None:
            start = index
        elif not missing and start is not None:
            runs.append((start, index - 1))
            start = None
    if start is not None:
        runs.append((start, len(mask) - 1))
    return runs


def impute_short_gaps(
    audit: pd.DataFrame,
) -> tuple[pd.DataFrame, list[list[Any]], pd.Series]:
    """Fill bounded gaps up to 12 hours while preserving structural gaps."""
    original_missing = audit[OUTPUT_COLUMNS[2:]].isna().sum()
    target_mask = (
        audit["_SOURCE_FILE"].eq("202602.xls")
        & audit["NTU"].isna()
    )
    summary: list[list[Any]] = []

    for column in NUMERIC_COLUMNS:
        filled = 0
        if column == "RIVER LEVEL":
            missing_before = int(audit[column].isna().sum())
            audit[column] = audit[column].interpolate(
                method="linear", limit_direction="both"
            ).round(2)
            filled = missing_before - int(audit[column].isna().sum())
            summary.append(
                [
                    column,
                    int(original_missing[column]),
                    filled,
                    int(audit[column].isna().sum()),
                    "全列线性插值（边界用最近有效值）",
                ]
            )
            continue
        protected = target_mask if column == "NTU" else pd.Series(
            False, index=audit.index
        )
        for start, end in missing_runs(audit[column].isna()):
            length = end - start + 1
            if (
                length > MAX_IMPUTE_GAP
                or protected.iloc[start : end + 1].any()
                or start == 0
                or end == len(audit) - 1
            ):
                continue
            before = audit.at[start - 1, column]
            after = audit.at[end + 1, column]
            if pd.isna(before) or pd.isna(after):
                continue
            values = np.linspace(float(before), float(after), length + 2)[1:-1]
            audit.loc[start:end, column] = values
            filled += length
        summary.append(
            [
                column,
                int(original_missing[column]),
                filled,
                int(audit[column].isna().sum()),
                "线性插值",
            ]
        )

    for column in PUMP_COLUMNS:
        filled = 0
        for start, end in missing_runs(audit[column].isna()):
            length = end - start + 1
            if length > MAX_IMPUTE_GAP or start == 0 or end == len(audit) - 1:
                continue
            before = audit.at[start - 1, column]
            after = audit.at[end + 1, column]
            if pd.isna(before) or pd.isna(after):
                continue
            context = pd.concat(
                [
                    audit.loc[max(0, start - 6) : start - 1, column],
                    audit.loc[end + 1 : min(len(audit) - 1, end + 6), column],
                ]
            ).dropna()
            modes = context.mode()
            if modes.empty:
                continue
            if len(modes) > 1 and before != after:
                continue
            value = before if before == after else modes.iloc[0]
            audit.loc[start:end, column] = value
            filled += length
        summary.append(
            [
                column,
                int(original_missing[column]),
                filled,
                int(audit[column].isna().sum()),
                "上下文众数",
            ]
        )

    return audit, summary, original_missing


def anomaly_record(
    frame: pd.DataFrame, index: int, field: str, value: Any
) -> dict[str, Any]:
    timestamp = frame.at[index, "_TIMESTAMP"]
    return {
        "timestamp": timestamp,
        "field": field,
        "value": value,
        "source": (
            f"{frame.at[index, '_SOURCE_FILE']}/"
            f"{frame.at[index, '_SOURCE_SHEET']}"
        ),
    }


def detect_hard_range_anomalies(audit: pd.DataFrame) -> list[dict[str, Any]]:
    anomalies: list[dict[str, Any]] = []
    for column, (lower, upper) in HARD_RANGES.items():
        values = audit[column]
        mask = values.notna() & (values < lower)
        if upper is not None:
            mask |= values.notna() & (values > upper)
        for index in audit.index[mask]:
            anomalies.append(anomaly_record(audit, index, column, values.at[index]))

    for column in ["R/W PUMP DUTY", "T/W PUMP DUTY"]:
        numeric = pd.to_numeric(audit[column], errors="coerce")
        mask = numeric.notna() & ((numeric < 0) | (numeric > 10))
        for index in audit.index[mask]:
            anomalies.append(anomaly_record(audit, index, column, audit.at[index, column]))
    return anomalies


def detect_hampel_outliers(
    audit: pd.DataFrame, window: int = 13, sigma: float = 5.0
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    counts: dict[str, int] = {}
    examples: list[dict[str, Any]] = []
    for column in NUMERIC_COLUMNS:
        series = audit[column].astype(float)
        if series.notna().sum() < window:
            counts[column] = 0
            continue
        median = series.rolling(window, center=True, min_periods=7).median()
        deviation = (series - median).abs()
        mad = deviation.rolling(window, center=True, min_periods=7).median()
        threshold = sigma * 1.4826 * mad
        mask = (
            series.notna()
            & median.notna()
            & mad.notna()
            & (mad > 0)
            & (deviation > threshold)
        )
        counts[column] = int(mask.sum())
        ranked = deviation[mask].sort_values(ascending=False).head(5)
        for index in ranked.index:
            item = anomaly_record(audit, index, column, series.at[index])
            item["local_median"] = median.at[index]
            item["deviation"] = deviation.at[index]
            examples.append(item)
    return counts, examples


def detect_exceedance_runs(audit: pd.DataFrame) -> list[dict[str, Any]]:
    mask = audit["NTU"].gt(1)
    runs: list[dict[str, Any]] = []
    start: int | None = None
    previous: int | None = None
    for index in audit.index[mask]:
        if previous is None or index != previous + 1:
            if start is not None and previous is not None:
                runs.append(build_run(audit, start, previous))
            start = index
        previous = index
    if start is not None and previous is not None:
        runs.append(build_run(audit, start, previous))
    return runs


def build_run(audit: pd.DataFrame, start: int, end: int) -> dict[str, Any]:
    values = audit.loc[start:end, "NTU"]
    return {
        "start": audit.at[start, "_TIMESTAMP"],
        "end": audit.at[end, "_TIMESTAMP"],
        "points": end - start + 1,
        "hours": (end - start + 1) * 2,
        "maximum": float(values.max()),
    }


def month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def format_value(value: Any) -> str:
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        cells = [str(value).replace("|", "\\|").replace("\n", " ") for value in row]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def build_report(
    audit: pd.DataFrame,
    repairs: list[str],
    pump_corrections: list[str],
    pump_source_counts: dict[str, Counter],
    imputation_summary: list[list[Any]],
    original_missing: pd.Series,
    invalid_numeric: list[dict[str, Any]],
    hard_anomalies: list[dict[str, Any]],
    outlier_counts: dict[str, int],
    outlier_examples: list[dict[str, Any]],
    duplicate_count: int,
) -> str:
    total = len(audit)
    months = sorted(audit["_SOURCE_DATE"].map(month_key).unique())
    target_mask = (
        audit["_SOURCE_FILE"].eq("202602.xls")
        & audit["NTU"].isna()
    )

    structural_masks: dict[str, pd.Series] = {}
    missing_rows: list[list[Any]] = []
    monthly_missing_rows: list[list[Any]] = []
    for column in OUTPUT_COLUMNS[2:]:
        missing = audit[column].isna()
        structural = pd.Series(False, index=audit.index)
        for month in months:
            month_mask = audit["_SOURCE_DATE"].map(month_key).eq(month)
            count = int((missing & month_mask).sum())
            month_total = int(month_mask.sum())
            if count == month_total:
                structural |= month_mask
            if count:
                monthly_missing_rows.append(
                    [month, column, count, f"{count / month_total:.1%}"]
                )
        structural_masks[column] = structural
        column_target = target_mask if column == "NTU" else pd.Series(
            False, index=audit.index
        )
        target_count = int(column_target.sum())
        ordinary = int((missing & ~structural & ~column_target).sum())
        missing_rows.append(
            [
                column,
                int(original_missing[column]),
                int(original_missing[column] - missing.sum()),
                int(missing.sum()),
                f"{missing.mean():.2%}",
                int((missing & structural & ~column_target).sum()),
                target_count,
                ordinary,
            ]
        )

    hard_combined = invalid_numeric + hard_anomalies
    hard_combined.sort(key=lambda item: (item["timestamp"], item["field"]))
    hard_rows = [
        [
            item["timestamp"].strftime("%Y-%m-%d %H:%M"),
            item["field"],
            format_value(item["value"]),
            item["source"],
        ]
        for item in hard_combined
    ]

    outlier_rows = [
        [column, count]
        for column, count in outlier_counts.items()
        if count
    ]
    example_rows = [
        [
            item["timestamp"].strftime("%Y-%m-%d %H:%M"),
            item["field"],
            format_value(item["value"]),
            format_value(item["local_median"]),
            item["source"],
        ]
        for item in sorted(
            outlier_examples,
            key=lambda item: item["deviation"],
            reverse=True,
        )[:30]
    ]

    exceedance = audit["NTU"].gt(1)
    exceedance_runs = detect_exceedance_runs(audit)
    run_rows = [
        [
            run["start"].strftime("%Y-%m-%d %H:%M"),
            run["end"].strftime("%Y-%m-%d %H:%M"),
            run["points"],
            run["hours"],
            f"{run['maximum']:.3g}",
        ]
        for run in sorted(
            exceedance_runs,
            key=lambda run: (run["points"], run["maximum"]),
            reverse=True,
        )[:30]
    ]

    repair_lines = "\n".join(f"- {item}" for item in repairs) or "- 无。"
    pump_correction_lines = (
        "\n".join(f"- {item}" for item in pump_corrections) or "- 无特殊纠错。"
    )
    pump_mapping_rows = []
    for column in PUMP_COLUMNS:
        for source_value, count in sorted(
            pump_source_counts[column].items(),
            key=lambda item: (-item[1], item[0]),
        ):
            if column == "T/W PUMP DUTY" and source_value in {"244", "2,2"}:
                converted_count = 2
            else:
                converted_count = len(re.findall(r"\d+", source_value))
            pump_mapping_rows.append(
                [column, source_value, converted_count, count]
            )
    date_note = (
        "2025 年月度附件中的日期混用 Excel 日期和 `日/月/年` 文本，"
        "且部分日期被软件按 `月/日` 显示。脚本依据文件月份、每日固定 12 行顺序"
        "重建运营日；`01:00/03:00/05:00` 归入下一自然日。"
    )

    return f"""{README_MARKER}
## 数据合并与质量审计

本节由 `temp/merge_and_audit.py` 自动生成。合并结果为
`data/merged.xlsx`，原始附件未被修改。

### 合并概况

- 数据量：{total} 条两小时观测。
- 时间范围：{audit["_TIMESTAMP"].min():%Y-%m-%d %H:%M} 至 {audit["_TIMESTAMP"].max():%Y-%m-%d %H:%M}。
- 时间戳重复：{duplicate_count} 条。
- 输出结构：1 个工作表、{len(OUTPUT_COLUMNS)} 个字段；已删除全空的 `18ML LEVEL`、`18ML FLOW` 和文本备注 `REMARKS`。
- 日期处理：{date_note}
- 2026 年 2 月 `NTU`：整月 {int(target_mask.sum())} 个预测目标均保持为空，未参与插补。

### 缺失值处理

- `RIVER LEVEL` 全部补齐：内部缺口采用线性插值，若缺失位于数据边界则使用最近有效值，结果保留两位小数。
- 连续缺失不超过 {MAX_IMPUTE_GAP} 个时点（12 小时）、且缺口两端都有有效值时，连续变量采用线性插值。
- 两个泵状态字段采用上下文众数，避免产生小数泵数量。
- 超过 12 小时的连续缺失保持不变。
- 2026 年 2 月全部 `NTU` 预测目标保持缺失。
- 插补值是建模用估计值，不属于原始实测数据。

{markdown_table(
    ["字段", "原始缺失", "已填补", "保留缺失", "方法"],
    sorted(imputation_summary, key=lambda row: OUTPUT_COLUMNS.index(row[0])),
)}

### 字段说明

字段含义依据题目附录整理。题目没有给出的物理单位不作推断。

{markdown_table(
    ["字段", "中文含义", "单位或备注"],
    [row for row in DATA_DICTIONARY if row[0] in OUTPUT_COLUMNS],
)}

### 字段缺失统计

“整月未提供”指某字段在某来源月份全部为空；“普通漏测”排除了整月未提供和预测目标留空。

{markdown_table(
    ["字段", "原始缺失", "已填补", "保留缺失", "保留缺失率", "整月未提供", "预测目标", "普通漏测"],
    missing_rows,
)}

### 按月份缺失明细

{markdown_table(["来源月份", "字段", "缺失数", "月内缺失率"], monthly_missing_rows)}

### 明确格式修复

{repair_lines}

### 水泵运行数量转换

原始 `R/W PUMP DUTY` 和 `T/W PUMP DUTY` 使用泵号组合表示运行状态，
分隔符包括 `+`、`,`、`&` 和 `/`。合并文件将两个字段统一改为**运行泵数量**：
提取单元格中的泵号数字，并按数字个数计数。
例如 `1+2+5` 转换为 3，`1,4` 转换为 2，裸值 `2` 转换为 1。

特殊纠错：

{pump_correction_lines}

全部原始写法及转换结果：

{markdown_table(["字段", "原始写法", "运行泵数量", "出现次数"], pump_mapping_rows)}

### 硬异常与非数字值

硬异常包括数值字段中的文本，以及明显违反量纲边界的值。它们仅被报告，除非法时间外未修改原值；
非数字值在数值型输出列中保存为空单元格。

{markdown_table(["时间", "字段", "原值", "来源"], hard_rows) if hard_rows else "未发现。"}

### 统计离群点

连续数值字段按时间顺序使用中心窗口 13 点（约 24 小时）的 Hampel/MAD 规则，
阈值为 `5 × 1.4826 × MAD`。离群点仅用于审计，不修改数值。局部 MAD 为 0 的窗口不判定。

{markdown_table(["字段", "离群点数"], outlier_rows) if outlier_rows else "未发现。"}

以下列出偏离程度最大的代表记录（最多 30 条）：

{markdown_table(["时间", "字段", "原值", "局部中位数", "来源"], example_rows) if example_rows else "无。"}

### 出厂水浊度超标

按题目给定国标，处理后水 `NTU > 1` 记为超标。该项属于水质风险，不等同于数据错误。

- 超标观测：{int(exceedance.sum())} 条，占有效 `NTU` 观测的 {exceedance.sum() / audit["NTU"].notna().sum():.2%}。
- 连续超标区间：{len(exceedance_runs)} 段。
- 最大 `NTU`：{audit["NTU"].max():g}。

持续时间最长的区间（最多 30 段）：

{markdown_table(["开始", "结束", "观测点", "持续小时", "区间最大 NTU"], run_rows) if run_rows else "无。"}

### 复现方法

```bash
conda env create -f temp/environment.yml
bash temp/run.sh
```
"""


def write_workbook(output: pd.DataFrame) -> None:
    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
        date_format="yyyy-mm-dd",
        datetime_format="yyyy-mm-dd hh:mm",
    ) as writer:
        output.to_excel(writer, sheet_name="Merged", index=False)

    workbook = load_workbook(OUTPUT_PATH)
    sheet = workbook["Merged"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 10
    for index in range(3, len(OUTPUT_COLUMNS) + 1):
        column = OUTPUT_COLUMNS[index - 1]
        width = max(11, min(20, len(column) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in sheet["B"][1:]:
        if isinstance(cell.value, str):
            cell.value = datetime.strptime(cell.value, "%H:%M:%S").time()
        cell.number_format = "hh:mm"
    river_level_column = OUTPUT_COLUMNS.index("RIVER LEVEL") + 1
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, river_level_column).number_format = "0.00"
    workbook.save(OUTPUT_PATH)


def update_readme(report: str) -> None:
    current = README_PATH.read_text(encoding="utf-8")
    if README_MARKER in current:
        current = current.split(README_MARKER, 1)[0].rstrip()
    README_PATH.write_text(current + "\n\n" + report.rstrip() + "\n", encoding="utf-8")


def verify_output(expected_rows: int) -> None:
    workbook = load_workbook(OUTPUT_PATH, read_only=False, data_only=True)
    if workbook.sheetnames != ["Merged"]:
        raise AssertionError(f"Unexpected sheets: {workbook.sheetnames}")
    sheet = workbook["Merged"]
    if sheet.max_row != expected_rows + 1:
        raise AssertionError(
            f"Expected {expected_rows + 1} worksheet rows, found {sheet.max_row}"
        )
    headers = [cell.value for cell in sheet[1]]
    if headers != OUTPUT_COLUMNS:
        raise AssertionError(f"Unexpected headers: {headers}")
    if sheet.freeze_panes != "A2":
        raise AssertionError("Freeze panes missing")
    if sheet.auto_filter.ref != sheet.dimensions:
        raise AssertionError("Auto filter does not cover the used range")
    if not isinstance(sheet["B2"].value, time):
        raise AssertionError(
            f"TIME cells are not native Excel times: {type(sheet['B2'].value)}"
        )
    for column in PUMP_COLUMNS:
        column_index = OUTPUT_COLUMNS.index(column) + 1
        pump_values = {
            sheet.cell(row, column_index).value
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, column_index).value is not None
        }
        if not pump_values.issubset({1, 2, 3}):
            raise AssertionError(
                f"Unexpected {column} values after conversion: {pump_values}"
            )


def main() -> None:
    records: list[dict[str, Any]] = []
    repairs: list[str] = []
    for path in sorted((DATA_DIR / "2025").glob("*.xlsx")):
        parsed, file_repairs = parse_2025_file(path)
        records.extend(parsed)
        repairs.extend(file_repairs)
    for path in sorted((DATA_DIR / "2026").glob("*.xls")):
        parsed, file_repairs = parse_2026_file(path)
        records.extend(parsed)
        repairs.extend(file_repairs)

    audit = pd.DataFrame.from_records(records)
    audit["_TIMESTAMP"] = pd.to_datetime(audit["DATE"].astype(str)) + pd.to_timedelta(
        audit["TIME"].map(lambda value: value.hour), unit="h"
    )
    audit = audit.sort_values("_TIMESTAMP", kind="stable").reset_index(drop=True)
    if len(audit) != 5460:
        raise AssertionError(f"Expected 5460 records, found {len(audit)}")
    duplicate_count = int(audit["_TIMESTAMP"].duplicated(keep=False).sum())

    audit, pump_corrections, pump_source_counts = convert_pump_duties(
        audit
    )
    audit, invalid_numeric = coerce_numeric_columns(audit)
    hard_anomalies = detect_hard_range_anomalies(audit)
    audit, imputation_summary, original_missing = impute_short_gaps(audit)
    outlier_counts, outlier_examples = detect_hampel_outliers(audit)

    output = audit[OUTPUT_COLUMNS].copy()
    write_workbook(output)
    report = build_report(
        audit,
        repairs,
        pump_corrections,
        pump_source_counts,
        imputation_summary,
        original_missing,
        invalid_numeric,
        hard_anomalies,
        outlier_counts,
        outlier_examples,
        duplicate_count,
    )
    update_readme(report)
    verify_output(len(output))

    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)} with {len(output)} rows.")
    print(f"Updated {README_PATH.relative_to(ROOT)}.")


if __name__ == "__main__":
    main()
